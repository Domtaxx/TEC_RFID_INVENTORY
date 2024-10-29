import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, and_
from fastapi.responses import StreamingResponse
from io import BytesIO
from database.session import get_db
import pandas as pd
from database.models import Item, Room, Department, ItemRegistry

router = APIRouter()

from sqlalchemy import func, and_
from sqlalchemy.orm import aliased

def get_items_by_department(department_id: int, db: Session):
    # Subquery to get the latest registry date for each item
    latest_registry_subquery = (
        db.query(
            ItemRegistry.id_item,
            func.max(ItemRegistry.registry_date).label("latest_registry_date")
        )
        .join(Room, ItemRegistry.id_room == Room.id)
        .filter(Room.id_department == department_id)  # Filter by department
        .group_by(ItemRegistry.id_item)
        .subquery()
    )

    # Main query to get the latest registry details for each item
    return (
        db.query(Item, ItemRegistry, Room, Department)
        .join(latest_registry_subquery, and_(
            Item.id == latest_registry_subquery.c.id_item
        ))
        .join(ItemRegistry, and_(
            ItemRegistry.id_item == latest_registry_subquery.c.id_item,
            ItemRegistry.registry_date == latest_registry_subquery.c.latest_registry_date
        ))
        .join(Room, ItemRegistry.id_room == Room.id)
        .join(Department, Room.id_department == Department.id)
        .filter(Room.id_department == department_id)  # Filter by department
        .all()
    )


@router.get("/items_by_department/{department_id}")
def generate_department_report(department_id: int, db: Session = Depends(get_db)):
    items = get_items_by_department(department_id, db)
    
    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items Report"

    # Add headers
    ws.append(["ID del Activo","Nombre del Activo", "Numero Serial", "Persona Encargada", "Nombre del departamento", "Ultima ubicacion conocida", "Dia registrado"])

    # Add data rows
    for item, registry, room, department in items:
        ws.append([item.id ,item.item_name, item.serial_number, department.department_name, room.room_name, registry.registry_date])

    # Save the file to a BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file as a response
    headers = {
        'Content-Disposition': f'attachment; filename=items_by_department_{department_id}.xlsx'
    }
    return StreamingResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.get("/items_by_room/{room_id}", response_class=StreamingResponse)
def generate_room_report(room_id: int, db: Session = Depends(get_db)):
    # Subquery to get the latest registry for each item
    subquery = db.query(
        ItemRegistry.id_item,
        func.max(ItemRegistry.registry_date).label('latest_registry_date')
    ).group_by(ItemRegistry.id_item).subquery()

    # Query to fetch the latest registered items in the specified room
    items = db.query(Item).join(ItemRegistry).join(Room).filter(
        and_(
            Room.id == room_id,
            ItemRegistry.id_item == subquery.c.id_item,
            ItemRegistry.registry_date == subquery.c.latest_registry_date
        )
    ).all()

    if not items:
        raise HTTPException(status_code=404, detail="No items found for the room.")

    # Create an Excel workbook and sheet using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Report"

    # Add the headers
    headers = ["ID del Activo","Nombre del Activo", "Numero Serial", "Persona Encargada", "Ultima ubicacion conocida", "Dia registrado"]
    ws.append(headers)

    # Add the data
    for item in items:
        latest_registry = max(item.item_registries, key=lambda x: x.registry_date)
        row = [
            item.id,
            item.item_name,
            item.serial_number,
            item.employee.first_name + " " + item.employee.surname,
            latest_registry.room.room_name,
            latest_registry.registry_date.strftime("%Y-%m-%d %H:%M:%S")
        ]
        ws.append(row)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the Excel file as a streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=room_report_{room_id}.xlsx"}
    )



@router.get("/items_by_employee/{employee_id}", response_class=StreamingResponse)
def generate_emp_report(employee_id: int, db: Session = Depends(get_db)):
     # Query to get the items registered by the employee with the latest registry information
    latest_registry_subquery = (
        db.query(
            ItemRegistry.id_item,
            func.max(ItemRegistry.registry_date).label("latest_registry_date")
        )
        .group_by(ItemRegistry.id_item)
        .subquery()
    )

    # Main query to fetch items, latest registry info, and room details
    items = (
        db.query(Item, ItemRegistry, Room)
        .join(ItemRegistry, Item.id == ItemRegistry.id_item)
        .join(Room, ItemRegistry.id_room == Room.id)
        .join(latest_registry_subquery, and_(
            latest_registry_subquery.c.id_item == ItemRegistry.id_item,
            latest_registry_subquery.c.latest_registry_date == ItemRegistry.registry_date
        ))
        .filter(Item.id_employee == employee_id)  # Filter by employee ID
        .all()
    )

    if not items:
        raise HTTPException(status_code=404, detail="No items found for the given employee.")

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Employee {employee_id} Items"

    # Write headers
    ws.append(["ID del Activo","Nombre del Activo", "Numero Serial", "Persona Encargada", "Ultima ubicacion conocida", "Dia registrado"])

    # Write item data
    for item, registry, room in items:
        latest_registry = max(item.item_registries, key=lambda x: x.registry_date)
        ws.append([
            item.id,
            item.item_name,
            item.serial_number,
            item.employee.first_name + " " + item.employee.surname,
            latest_registry.room.room_name,
            latest_registry.registry_date.strftime("%Y-%m-%d %H:%M:%S")
        ])

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the Excel file as a response
    headers = {
        'Content-Disposition': f'attachment; filename="employee_{employee_id}_items.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)